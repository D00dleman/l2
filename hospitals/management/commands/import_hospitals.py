from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from hospitals.models import Hospitals


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с медорганизациями
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        print(ws)  # noqa: T001
        starts = False
        code, full_title, short_title, address = '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код" in cells and "краткое" in cells and "полное" in cells and "адрес" in cells:
                    starts = True
                    code = cells.index("код")
                    short_title = cells.index("краткое")
                    full_title = cells.index("полное")
                    address = cells.index("адрес")
            else:
                if Hospitals.objects.filter(code_tfoms=cells[code]).exists():
                    print(Hospitals.objects.filter(code_tfoms=cells[code]))  # noqa: T001
                Hospitals.objects.create(title=cells[full_title][:255], short_title=cells[short_title][:255], code_tfoms=cells[code], address=cells[address][:128])
                print(f'добавлено МО:{cells[code]}:{cells[full_title]}:{cells[short_title]}:{cells[address]}:')  # noqa: T001
